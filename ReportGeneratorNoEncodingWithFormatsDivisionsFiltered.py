#!/usr/bin/env python
# coding: utf-8

# In[1]:


import cx_Oracle
import numpy as np
import pandas as pd
import datetime as dt
import openpyxl
from openpyxl.styles.fills import PatternFill
from openpyxl.utils.cell import get_column_letter

date = input('Введите дату в формате ГГГГ-ММ-ДД и нажмите ENTER:') # сюда подставляем нужную дату в формате date_format_python
date_format_python = '%Y-%m-%d' # если удобен другой формат, поменяйте на свой, код отработает нормально

date_format_sql = 'yyyy-mm-dd' # это формат для sql

try: 
    dt.datetime.strptime(date, date_format_python)
except:
    print('Неверный формат даты, будут показаны данные за сегодня, если они есть')
    date = str(dt.date.today())

con = cx_Oracle.connect('nbln/jnltkppo@192.168.171.109:1521/elev') # здесь в скобки надо поставить ваш путь соединения в формате 'user/password@dsn'
cur = con.cursor()
query = f"""
SELECT 
DIVISION.NAME,DIVISION.SNAME AS sname,FILIAL.ALIAS,FILIAL.ELEV,
CULTURE.name AS culturename,OSTATKI.W_XRA AS remnants, OSTATKI.recipient
FROM OSTATKI
INNER JOIN DIVISION
ON OSTATKI.division = DIVISION.NUM
INNER JOIN FILIAL
ON OSTATKI.division = FILIAL.division
INNER JOIN CULTURE
ON OSTATKI.culture = CULTURE.NUM
WHERE OSTATKI.sutki = TO_DATE('{date}','{date_format_sql}') 
AND OSTATKI.culture <= 9999 
AND OSTATKI.W_XRA IS NOT null
AND OSTATKI.DIVISION NOT IN (1,40,50) 
"""
cur.execute(query)
result = cur.fetchall()
cur.close()
con.close()

dtable = pd.DataFrame(result, columns = ['NAME', 'SNAME', 'ALIAS', 'ELEV', 'CULTURENAME', 'REMNANTS',
       'RECIPIENT'])

def getfilial(x,y):
    if x == y:
        return x
    else:
        return x + " " + y

def categorizerecipients(x):
    if x == 0:
        return 0
    else:
        return 1

# create supplementary fields
# merge name and sname if necessary
dtable['філіі'] = [*map(lambda x,y: getfilial(x,y),dtable['NAME'],dtable['SNAME'])]
dtable['RECIPIENT'] = [*map(lambda x: categorizerecipients(x),dtable['RECIPIENT'])]

# for pivot with total sums
dtable['Future_index'] = 2

# cross_table
cross = pd.crosstab([dtable['філіі'],dtable['ALIAS'],dtable['ELEV'],dtable['Future_index']], [dtable['CULTURENAME'],dtable['RECIPIENT']], dtable['REMNANTS'], aggfunc = 'sum')

# cross_table 2 --- to add future index
df2 = cross.sum(level=[0], axis=1)
df2 = df2.unstack(level=-1)
df2.columns = df2.columns.rename("RECIPIENT", level=1)

# delete index to merge with df2
cross = cross.reset_index(level='Future_index')
cross = cross.drop(columns='Future_index')

# unite cross and df2
result = pd.concat([cross, df2], axis=1, sort=True)
result = result.sort_index(axis=1, level=0, ascending=True, inplace=False, kind='quicksort', na_position='last', sort_remaining=True, by=None)

# formatting
result = result.rename({0:'Своє',1:'інші',2:'разом'},level=1, axis = 1)
result.columns.names = ['Культура', 'Залишки']

# estimate column sums
result_for_selection = result.reset_index(level=['ALIAS','ELEV'])

selection1 = result_for_selection[(result_for_selection['ELEV'].isin([16,4]) 
                                   & (result_for_selection['ALIAS'] != 'STAR'))
                                   & (~result_for_selection['ALIAS'].str.contains('^F', regex=True))]
itogo1 = 'Всього по автотранспорту:'

selection2 = result_for_selection[result_for_selection['ELEV'] == 2]
itogo2 = 'Всього по флоту СК:'

selection3 = result_for_selection[(result_for_selection['ELEV'] == 1) | (result_for_selection['ALIAS'] == 'STAR')]
itogo3 = 'Всього по залізничному транспорту:'

selections_list = [[selection1,itogo1],[selection2,itogo2],[selection3,itogo3]]

# calculate all sums

final_df = pd.DataFrame()
for combo in selections_list:
    selection = combo[0]
    itogo = combo[1]
    selection.loc[(itogo), selection.columns] = selection.sum()
    final_df = pd.concat([final_df,selection])

itogo4 = 'Загальна кількість по всім філіям:'

# delete supplementary columns
final_df = final_df.drop(columns = ['ALIAS','ELEV'])

final_df.loc[(itogo4), final_df.columns] = (sum([final_df.loc[itogo1],final_df.loc[itogo2],final_df.loc[itogo3]]))

# configure display options
pd.options.display.max_columns = len(final_df.columns)
pd.options.display.max_rows = len(final_df)

final_df.to_excel(f'Otchet{date}.xlsx')

wb = openpyxl.load_workbook(f'Otchet{date}.xlsx')
sheet = wb.active

def as_text(x):
    if x == None:
        return ''
    else:
        return str(x)
    
for column_cells in sheet.iter_cols():
    length = max(len(as_text(cell.value)) for cell in column_cells)
    letter = get_column_letter(column_cells[0].column)
    sheet.column_dimensions[letter].width = length + 5

for row in sheet.iter_rows():
    if row[0].value in ['Всього по автотранспорту:','Всього по флоту СК:','Всього по залізничному транспорту:']:
        for cell in row:
            cell.fill = PatternFill(start_color="FFF633", end_color="FFF633", fill_type='solid')
    elif row[0].value == 'Загальна кількість по всім філіям:':
        for cell in row:
            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type='solid')


wb.save(f'Otchet{date}.xlsx')

